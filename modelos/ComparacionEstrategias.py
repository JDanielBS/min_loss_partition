import math
import itertools
import os

from modelos.AlgoritmoPrincipal import AlgoritmoPrincipal
from modelos.AlgoritmoFuerzaBruta import AlgoritmoFuerzaBruta
from modelos.PyphiCode import PyphiCode
from modelos.sistema import Sistema
from itertools import combinations
from openpyxl import load_workbook
from icecream import ic
import pandas as pd

sistemaCambiante = Sistema('100010', '111111', '111111', '111111')

class ComparacionEstrategias:
    
    def __init__(self, ruta_e_n, ruta_e_e):
        self.ruta_e_n = ruta_e_n
        self.ruta_e_e = ruta_e_e
        self.combinaciones_e3_fb = []
        self.combinaciones_e3_pyphi = []

    """
    ------------------------------------------------------------------------------------------------
    Generacion subsistemas
    ------------------------------------------------------------------------------------------------
    """
    def todos_los_subsistemas(self, n_bits):
        """
        Genera todas las combinaciones posibles de dos números de 6 bits con al menos un 1.

        Args:
            n_bits (int): El número de bits de los números binarios.
        
        Returns:
            list: Una lista de tuplas con los números binarios generados.
        """
        # Paso 1: Representar los números binarios
        bits = n_bits  # Número de bits
        options = [[0, 1] for _ in range(bits)]  # Cada bit puede ser 0 o 1

        # Paso 2: Generar todas las combinaciones posibles
        combinations = [
            (num1, num2)
            for num1 in itertools.product(*options)
            for num2 in itertools.product(*options)
            if (
                sum(num1) + sum(num2) >= 3 and  # La suma total de unos en ambos números es 3
                sum(num1) >= 1 and sum(num2) >= 1  # Cada número debe tener al menos un 1
            )
        ]

        # Paso 3: Combinar los números en una sola estructura (cadenas concatenadas)
        merged_combinations = [
            (''.join(map(str, num1)), ''.join(map(str, num2)))
            for num1, num2 in combinations
        ]

        return merged_combinations

    def generar_combinaciones_binarias(self, n, max_combinaciones=50):
        """
        Genera todas las combinaciones posibles de las posiciones de un 0 en un número binario
        de tamaño 2n lleno de unos.

        Args:
            n (int): El número de bits del número binario.
            max_combinaciones (int): El número máximo de combinaciones a generar.
        
        Returns:
            list: Una lista de tuplas con los números binarios generados.
        """
        resultado_final = []  # Lista para almacenar todas las tuplas generadas

        for ceros in range(1, n + 1):
            binario_inicial = "1" * (2 * n)  # Número binario inicial lleno de unos
            posiciones = list(range(len(binario_inicial)))  # Índices de las posiciones
            combinaciones = combinations(posiciones, ceros)  # Generar combinaciones de índices para los ceros

            # Agregar la combinación inicial llena de unos
            mitad = len(binario_inicial) // 2
            tupla_inicial = (binario_inicial[:mitad], binario_inicial[mitad:])
            resultado_final.append(tupla_inicial)

            for i, combinacion in enumerate(combinaciones):
                if i >= max_combinaciones - 1:  # Restar 1 porque ya incluimos la inicial
                    break  # Detener si se alcanza el límite de combinaciones

                # Crear el binario con los ceros en las posiciones de la combinación
                lista_binaria = list(binario_inicial)
                for idx in combinacion:
                    lista_binaria[idx] = '0'
                
                # Dividir la cadena binaria a la mitad y agregarla como tupla
                tupla_binaria = ("".join(lista_binaria[:mitad]), "".join(lista_binaria[mitad:]))
                resultado_final.append(tupla_binaria)

        return resultado_final

    """
    ------------------------------------------------------------------------------------------------
    Funciones auxiliares
    ------------------------------------------------------------------------------------------------
    """
    def obtener_indices(self, subsistema):
        """
        Traduce un subsistema binarios a los indices de los bits que son 1.

        Args:
            subsistema (str): El subsistema en formato binario.
        
        Returns:
            tuple: Los índices de los bits que son 1.
        """
        posiciones = [i for i, bit in enumerate(subsistema) if bit == '1']
        return tuple(posiciones)
    
    def convertir_e_inicial(self, estado_inicial):
        """
        Convierte el estado inicial a una tupla de enteros.

        Args:
            estado_inicial (str): El estado inicial en formato binario.
        
        Returns:
            tuple: Una tupla de enteros con los valores del estado inicial.
        """
        return tuple(int(bit) for bit in estado_inicial)

    def son_cercanos(self, num1, num2, tol=1e-9):
        """
        Determina la cercanía entre dos números (Son cercanos si sus 9 primeros decimales coinciden).

        Args:
            num1 (float): El primer número a comparar.
            num2 (float): El segundo número a comparar.
            tol (float): La tolerancia para determinar si los números son cercanos.
        
        Returns:
            bool: True si los números son cercanos, False en caso contrario.
        """
        return math.isclose(num1, num2, rel_tol=tol, abs_tol=tol)

    """
    ------------------------------------------------------------------------------------------------
    Estrategias
    ------------------------------------------------------------------------------------------------
    """
    def estrategia3(self, sistema):
        """
        Realiza la estrategia 3.

        Args:
            sistema (Sistema): El sistema (estado inicial, sistema candidato, 
            subsistema futuro y subsistema presente) a analizar.

        Returns:
            dict: Un diccionario con la solución de la estrategia 3.
        """
        algoritmo3 = AlgoritmoPrincipal(self.ruta_e_e, sistema)
        return algoritmo3.estrategia3()

    def fuerza_bruta(self, sistema):
        """
        Realiza la estrategia de fuerza bruta.

        Args:
            sistema (Sistema): El sistema (estado inicial, sistema candidato, 
            subsistema futuro y subsistema presente) a analizar.
        
        Returns:
            list: Lista de soluciones encontradas por fuerza bruta.
        """
        algoritmo0 = AlgoritmoFuerzaBruta(self.ruta_e_e, sistema)
        return algoritmo0.estrategia_fuerza_bruta()

    def pyphi_code_15(self, purview_e, mechanism_e):
        labels = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O')
        i_state = (1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        c_system = (0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14)
        purview = self.obtener_indices(purview_e)
        mechanism = self.obtener_indices(mechanism_e)
        pyphi_code = PyphiCode(archivo_estado_nodo, labels, i_state, c_system)
        return pyphi_code.pyphi_partition(mechanism, purview)

    def pyphi_code(self, pyphi_info):
        """
        Realiza el código de pyphi.

        Args:
            pyphi_info (dict): Un diccionario con la información necesaria para ejecutar pyphi.
            La información contiene el estado inicial, el sistema candidato, los labels,
            el subsistema futuro y el subsistema presente.
        
        Returns:
            dict: Un objeto con la solución de pyphi.
        """
        purview = self.obtener_indices(pyphi_info['purview'])
        mechanism = self.obtener_indices(pyphi_info['mechanism'])
        pyphi_code = PyphiCode(self.ruta_e_n, pyphi_info['labels'], pyphi_info['i_state'], pyphi_info['c_system'])
        return pyphi_code.pyphi_partition(mechanism, purview)

    """
    ------------------------------------------------------------------------------------------------
    Pruebas estrategia 3
    ------------------------------------------------------------------------------------------------
    """
    def encontrar_particion_igual_e3(self, listado_soluciones_fb, solucion_e3):
        """
        Compara la partición encontrada por la estrategia 3 con las particiones encontradas por fuerza bruta
        para averiguar si existe una partición igual.

        Args:
            listado_soluciones_fb (list): Lista de soluciones encontradas por fuerza bruta.
            solucion_e3 (dict): Un dicccionario con la solución de la estrategia 3.
        
        Returns:
            boolean: True si existe una partición igual, False en caso contrario.
            dict: Un diccionario con la partición igual encontrada.
        """
        particion = set(solucion_e3['particion1'])
        existe_solucion = False
        solucion_fb = None

        for solucion in listado_soluciones_fb:
            if (particion == set(solucion['particion1'])) or (particion == set(solucion['particion2'])):
                existe_solucion = True
                solucion_fb = solucion
                break

        return existe_solucion, solucion_fb

    def intentos_con_combinaciones_e3(self):
        """
        Realiza intentos con todas las combinaciones de subsistemas posibles.
        """
        subsistemas = todos_los_subsistemas(6)

        for i, (num1, num2) in enumerate(subsistemas):
            sistemaCambiante.set_subsistema_futuro(num1)
            sistemaCambiante.set_subsistema_presente(num2)

            listado_soluciones_fb, tiempo_fb = fuerza_bruta(sistemaCambiante)
            existe_solucion = False

            while not existe_solucion:
                solucion_e3, tiempo_e3 = self.estrategia3(sistemaCambiante)
                existe_solucion, solucion_fb = encontrar_particion_igual_e3(listado_soluciones_fb, solucion_e3)
            
            mejor_particion = {
                'subsistema_futuro': num1,
                'subsistema_presente': num2,
                'particion_fb': str(set(solucion_fb['particion1'])) + ' - y - ' + str(set(solucion_fb['particion2'])),
                'perdida_fb': solucion_fb['emd'],
                'tiempo_fb': tiempo_fb,
                'particion_e3': str(set(solucion_e3['particion1'])) + ' - y - ' + str(set(solucion_e3['particion2'])),
                'perdida_e3': solucion_e3['emd'],
                'tiempo_e3': tiempo_e3
            }

            self.combinaciones_e3_fb.append(mejor_particion)

    """
    ------------------------------------------------------------------------------------------------
    Pruebas pyphi y estrategia 3
    ------------------------------------------------------------------------------------------------
    """
    def intentos_con_combinaciones_pyphi(self):
        """
        Realiza intentos con varias combinaciones de subsistemas casi completos, el número de pruebas
        depende del subconjunto de subsistemas generados.
        Las pruebas se pueden manejar mediante la variable subconjunto_subsistemas.
        """
        subsistemas = self.generar_combinaciones_binarias(15)
        subconjunto_subsistemas = [subsistemas[172]]
        ic(subconjunto_subsistemas)

        for i, (num1, num2) in enumerate(subconjunto_subsistemas):
            num1, num2 = str(num1), str(num2)
            sistemaCambiante.set_subsistema_futuro(num1)
            sistemaCambiante.set_subsistema_presente(num2)

            solucion_pyphi, tiempo_pyphi = self.pyphi_code_15(num1, num2)
            ic(solucion_pyphi['emd'])
            existe_solucion = False
            intentos = 0

            while not existe_solucion:
                intentos += 1
                solucion_e3, tiempo_e3, tiempo_e3_no_m = self.estrategia3(sistemaCambiante)
                existe_solucion = self.son_cercanos(solucion_pyphi['emd'], solucion_e3['emd'])
                ic(existe_solucion, solucion_pyphi['emd'], solucion_e3['emd'], str(solucion_pyphi['min_info_part']), solucion_e3['particion_formateada'])
            
            mejor_particion = {
                'subsistema_futuro': num1,
                'subsistema_presente': num2,
                'e1': ' ',
                'particion_pyphi': str([solucion_pyphi['prim_purv'], solucion_pyphi['prim_mech']]) + ' - y - ' + str([solucion_pyphi['dual_purv'], solucion_pyphi['dual_mech']]),
                'particion_pyphi_f': str(solucion_pyphi['min_info_part']),
                'perdida_pyphi': solucion_pyphi['emd'],
                'dist_pyphi': solucion_pyphi['distribution'],
                'tiempo_pyphi': tiempo_pyphi,
                'e2': ' ',
                'particion_e3': str(set(solucion_e3['particion1'])) + ' - y - ' + str(set(solucion_e3['particion2'])),
                'particion_e3_f': solucion_e3['particion_formateada'],
                'perdida_e3': solucion_e3['emd'],
                'dist_e3': solucion_e3['distribucion_experimental'],
                'tiempo_e3': tiempo_e3,
                'tiempo_e3_no_m': tiempo_e3_no_m,
                'intentos': intentos
            }

            self.combinaciones_e3_pyphi.append(mejor_particion)

    def prueba_con_subsistema(self, subsistema, info_pyphi):
        """
        Realiza un prueba con un único subsistema.

        Args:
            subsistema (tuple): El subsistema a analizar.
            info_pyphi (dict): Un diccionario con la información necesaria para ejecutar pyphi.
            La información contiene el estado inicial, el sistema candidato, los labels,
            el subsistema futuro y el subsistema presente.
        """
        purview, mechanism = str(subsistema.get_subsistema_futuro()), str(subsistema.get_subsistema_presente())

        solucion_pyphi, tiempo_pyphi = self.pyphi_code(info_pyphi)
        existe_solucion = False
        intentos = 0

        while not existe_solucion:
            intentos += 1
            solucion_e3, tiempo_e3 = self.estrategia3(sistemaCambiante)
            existe_solucion = self.son_cercanos(solucion_pyphi['emd'], solucion_e3['emd'])
            # ic(existe_solucion, solucion_pyphi['emd'], solucion_e3['emd'], str(solucion_pyphi['min_info_part']), solucion_e3['particion_formateada'])
            
        mejor_particion = {
            'subsistema_futuro': purview,
            'subsistema_presente': mechanism,
            'e1': ' ',
            'particion_pyphi': str([solucion_pyphi['prim_purv'], solucion_pyphi['prim_mech']]) + ' - y - ' + str([solucion_pyphi['dual_purv'], solucion_pyphi['dual_mech']]),
            'particion_pyphi_f': str(solucion_pyphi['min_info_part']),
            'perdida_pyphi': solucion_pyphi['emd'],
            'dist_pyphi': solucion_pyphi['distribution'],
            'tiempo_pyphi': tiempo_pyphi,
            'e2': ' ',
            'particion_e3': str(set(solucion_e3['particion1'])) + ' - y - ' + str(set(solucion_e3['particion2'])),
            'particion_e3_f': solucion_e3['particion_formateada'],
            'perdida_e3': solucion_e3['emd'],
            'dist_e3': solucion_e3['distribucion_experimental'],
            'tiempo_e3': tiempo_e3,
            'intentos': intentos
        }

        self.combinaciones_e3_pyphi.append(mejor_particion)

    """
    ------------------------------------------------------------------------------------------------
    Pruebas - guardar en excel
    ------------------------------------------------------------------------------------------------
    """
    def pasar_a_excel_e3_fb(self):

        self.intentos_con_combinaciones_e3()

        # Crear un DataFrame a partir de la lista de JSON
        df = pd.DataFrame(self.combinaciones_e3_fb)

        # Escribir el DataFrame en un archivo Excel
        nombre_archivo = 'solucion_e3_fb.xlsx'
        df.to_excel(nombre_archivo, index=False)

    def pasar_a_excel_e3_pyphi(self):
        
        self.intentos_con_combinaciones_pyphi()

        # Crear un DataFrame a partir de la lista de JSON
        df = pd.DataFrame(self.combinaciones_e3_pyphi)

        # Escribir el DataFrame en un archivo Excel
        nombre_archivo = 'solucion_pyphi.xlsx'
        df.to_excel(nombre_archivo, index=False)

    def minima_perdida(self, subsistema, info_pyphi):
        self.prueba_con_subsistema(subsistema, info_pyphi)

        # Crear un DataFrame a partir de la lista de JSON
        df = pd.DataFrame(self.combinaciones_e3_pyphi)

        # Convertir listas a cadenas en el DataFrame
        df = df.applymap(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)

        # Nombre del archivo Excel
        nombre_archivo = 'solucion_pyphi.xlsx'

        # Verificar si el archivo existe
        if not os.path.exists(nombre_archivo):
            # Si no existe, crear el archivo con los datos del DataFrame
            df.to_excel(nombre_archivo, index=False)
            print(f"Archivo '{nombre_archivo}' creado con éxito.")
        else:
            # Si existe, cargar el archivo
            workbook = load_workbook(nombre_archivo)
            sheet = workbook.active

            # Obtener los encabezados actuales del archivo
            headers_excel = [cell.value for cell in sheet[1]]  # Leer la primera fila (encabezados)
            headers_df = list(df.columns)  # Encabezados del DataFrame

            # Verificar si los encabezados coinciden
            if headers_excel != headers_df:
                raise ValueError("Los encabezados del DataFrame no coinciden con los del archivo Excel.")

            # Encontrar la última fila disponible
            last_row = sheet.max_row

            # Escribir los datos del DataFrame en el archivo Excel
            for index, row in df.iterrows():
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=last_row + 1 + index, column=col_index, value=value)

            # Guardar los cambios en el archivo Excel
            workbook.save(nombre_archivo)
            print(f"Datos agregados al archivo '{nombre_archivo}'.")